"""엑셀 내보내기 (openpyxl 직접 사용, pandas 불필요).

구버전은 ``df.to_excel`` 한 줄이라 서식이 전혀 없었다.
여기서는 헤더 강조 / 열 너비 / 통화 서식 / 틀 고정 / 자동 필터 / 요약 시트까지 만든다.
"""

from __future__ import annotations

import logging
from collections.abc import Sequence
from pathlib import Path

from chugui.models import Guest
from chugui.services.messages import MessageService
from chugui.services.settlement import Settlement

logger = logging.getLogger(__name__)

_HEADERS: tuple[tuple[str, int], ...] = (
    ("순번", 6),
    ("성명", 18),
    ("축의금액", 14),
    ("관계분류", 14),
    ("소속", 18),
    ("참석여부", 12),
    ("수령경로", 12),
    ("대인식권", 10),
    ("소인식권", 10),
    ("비고", 20),
    ("확인필요", 22),
    ("감사메시지", 60),
    ("발송완료", 12),
)


def export_to_excel(
    file_path: str | Path,
    guests: Sequence[Guest],
    settlement: Settlement,
    message_service: MessageService | None = None,
) -> Path:
    """하객 명단과 정산 요약을 xlsx로 저장한다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    service = message_service or MessageService()
    path = Path(file_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "축의금 명단"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    center = Alignment(horizontal="center", vertical="center")

    for column, (title, width) in enumerate(_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        sheet.column_dimensions[get_column_letter(column)].width = width

    for row_index, guest in enumerate(guests, start=2):
        values = (
            guest.guest_id,
            guest.name,
            guest.amount,
            guest.relation.value,
            guest.belong,
            guest.attendance.value,
            guest.payment.value,
            guest.adult_tickets,
            guest.child_tickets,
            guest.note,
            " / ".join(guest.warnings),
            service.generate(guest),
            "완료" if guest.sent_thanks else "미발송",
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            if column == 3:
                cell.number_format = "#,##0"
            elif column in (1, 6, 7, 8, 9, 13):
                cell.alignment = center
        if guest.needs_review:
            for column in range(1, len(_HEADERS) + 1):
                sheet.cell(row=row_index, column=column).fill = PatternFill("solid", fgColor="FEF3C7")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{max(1, len(guests) + 1)}"

    _write_summary_sheet(workbook, settlement)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    logger.info("엑셀 내보내기 완료: %s (%d건)", path, len(guests))
    return path


def _write_summary_sheet(workbook: object, settlement: Settlement) -> None:
    from openpyxl.styles import Alignment, Font

    sheet = workbook.create_sheet("정산 요약")  # type: ignore[attr-defined]
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 20

    rows: tuple[tuple[str, object], ...] = (
        ("총 수령 축의금", settlement.total_amount),
        ("총 하객 수(건)", settlement.guest_count),
        ("총 인원 수(명)", settlement.head_count),
        ("참석", settlement.attendee_count),
        ("불참(송금)", settlement.absentee_count),
        ("현금 합계", settlement.cash_amount),
        ("계좌이체 합계", settlement.transfer_amount),
        ("대인 식권 수", settlement.adult_tickets),
        ("대인 단가", settlement.adult_unit_cost),
        ("소인 식권 수", settlement.child_tickets),
        ("소인 단가", settlement.child_unit_cost),
        ("총 식대", settlement.meal_cost),
        ("최종 순 정산금", settlement.net_amount),
        ("확인 필요 건수", settlement.review_count),
        ("감사 인사 발송 완료", settlement.sent_count),
    )

    title = sheet.cell(row=1, column=1, value="정산 요약")
    title.font = Font(bold=True, size=14)

    for offset, (label, value) in enumerate(rows, start=3):
        label_cell = sheet.cell(row=offset, column=1, value=label)
        label_cell.font = Font(bold=True)
        value_cell = sheet.cell(row=offset, column=2, value=value)
        value_cell.alignment = Alignment(horizontal="right")
        if isinstance(value, int) and "수" not in label and "건수" not in label:
            value_cell.number_format = "#,##0"
