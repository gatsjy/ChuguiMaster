"""스프레드시트(.xlsx) / CSV → Guest 변환.

pandas를 쓰지 않는다. 이 프로그램이 pandas에서 쓰던 기능은 ``read_excel`` /
``to_excel`` 두 개뿐인데, 그 대가로 PyInstaller 산출물이 두 배가 되고
콜드 스타트가 몇 초씩 늘어난다. openpyxl + 표준 ``csv`` 로 충분하다.

구버전 대비 실질적인 개선 두 가지:

* **헤더 행 자동 탐색** - 은행/토스/카뱅 거래내역 엑셀은 상단에 계좌·기간
  안내가 5~10줄 붙어 나온다. 구버전은 그 첫 줄을 헤더로 잡아 통째로 실패했다.
* **CSV 실제 지원** - 구버전은 파일 다이얼로그와 드래그&드롭에서 ``*.csv`` 를
  받아놓고 ``pd.read_excel`` 만 호출해 ``ValueError`` 로 죽었다.
"""

from __future__ import annotations

import csv
import logging
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from chugui.models import (
    WARN_NO_AMOUNT,
    WARN_NO_NAME,
    Attendance,
    Guest,
    Payment,
    Source,
    renumber,
)
from chugui.parsing.amount import parse_amount
from chugui.parsing.names import format_display_name, split_names
from chugui.parsing.relations import guess_relation

logger = logging.getLogger(__name__)

SUPPORTED_SUFFIXES: tuple[str, ...] = (".xlsx", ".xlsm", ".csv")

# 헤더 후보 키워드. 은행 거래내역 컬럼명까지 포함한다.
_NAME_KEYS = ("성명", "이름", "입금자", "보낸분", "보내는분", "거래자", "성함", "고객명", "적요")
_AMOUNT_KEYS = ("금액", "축의금", "입금액", "입금", "받은금액", "거래금액", "원화금액")
_BELONG_KEYS = ("소속", "관계", "그룹", "분류", "구분")
_NOTE_KEYS = ("비고", "특이사항", "메모", "내용", "기재내용", "적요")
_TICKET_ADULT_KEYS = ("식권", "대인", "성인")
_TICKET_CHILD_KEYS = ("소인", "어린이", "아동")

# 헤더 탐색 시 스캔할 최대 행 수.
_MAX_HEADER_SCAN = 20


class ExcelParseError(RuntimeError):
    """스프레드시트를 해석할 수 없을 때."""


def _norm(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _match_column(headers: Sequence[str], keys: Sequence[str]) -> int | None:
    """키워드에 가장 잘 맞는 컬럼 인덱스. 긴 키워드가 우선."""
    best_index: int | None = None
    best_length = 0
    for index, header in enumerate(headers):
        for key in keys:
            if key in header and len(key) > best_length:
                best_index, best_length = index, len(key)
    return best_index


def _score_header_row(row: Sequence[str]) -> int:
    """이 행이 헤더일 가능성 점수."""
    filled = [cell for cell in row if cell]
    if len(filled) < 2:
        return 0
    score = 0
    if _match_column(row, _NAME_KEYS) is not None:
        score += 2
    if _match_column(row, _AMOUNT_KEYS) is not None:
        score += 2
    if _match_column(row, _BELONG_KEYS) is not None:
        score += 1
    if _match_column(row, _NOTE_KEYS) is not None:
        score += 1
    return score


def _find_header_row(rows: Sequence[Sequence[str]]) -> int:
    """상단 안내문을 건너뛰고 실제 헤더 행 인덱스를 찾는다."""
    best_index, best_score = 0, 0
    for index, row in enumerate(rows[:_MAX_HEADER_SCAN]):
        score = _score_header_row(row)
        if score > best_score:
            best_index, best_score = index, score
    if best_score < 2:
        logger.warning("헤더 행을 확신할 수 없어 첫 행을 헤더로 사용합니다.")
        return 0
    return best_index


# --------------------------------------------------------------------- 로더


def _read_csv_rows(path: Path) -> list[list[str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.read(4096)
                handle.seek(0)
                try:
                    dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                return [[_norm(cell) for cell in row] for row in csv.reader(handle, dialect)]
        except (UnicodeDecodeError, LookupError) as exc:
            last_error = exc
            continue
    raise ExcelParseError(f"CSV 인코딩을 판별하지 못했습니다: {last_error}")


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 배포본에는 항상 포함
        raise ExcelParseError("openpyxl이 설치되어 있지 않습니다.") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"엑셀 파일을 열 수 없습니다: {exc}") from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise ExcelParseError("시트를 찾을 수 없습니다.")
        return [[_norm(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _load_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path)
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path)
    if suffix == ".xls":
        raise ExcelParseError(
            "구형 .xls 형식은 지원하지 않습니다. 엑셀에서 .xlsx로 저장한 뒤 다시 시도해 주세요."
        )
    raise ExcelParseError(f"지원하지 않는 형식입니다: {suffix or '(확장자 없음)'}")


# -------------------------------------------------------------------- 파서


def _cell(row: Sequence[str], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return row[index]


def _int_cell(row: Sequence[str], index: int | None) -> int:
    text = _cell(row, index)
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else 0


def parse_rows(rows: Sequence[Sequence[str]], source: Source = Source.EXCEL) -> list[Guest]:
    """이미 읽어들인 행 목록을 Guest로 변환한다(테스트 진입점)."""
    if not rows:
        return []

    header_index = _find_header_row(rows)
    headers = list(rows[header_index])
    body: Iterable[Sequence[str]] = rows[header_index + 1 :]

    name_col = _match_column(headers, _NAME_KEYS)
    amount_col = _match_column(headers, _AMOUNT_KEYS)
    belong_col = _match_column(headers, _BELONG_KEYS)
    note_col = _match_column(headers, _NOTE_KEYS)
    adult_col = _match_column(headers, _TICKET_ADULT_KEYS)
    child_col = _match_column(headers, _TICKET_CHILD_KEYS)

    if name_col is None:
        name_col = 0
    if name_col == note_col:  # '적요'가 이름과 비고 양쪽에 걸린 경우
        note_col = None

    guests: list[Guest] = []
    for offset, row in enumerate(body, start=1):
        if not any(cell for cell in row):
            continue

        raw_name = _cell(row, name_col)
        if not raw_name or raw_name in ("nan", "None"):
            continue
        # 은행 엑셀 하단의 합계/소계 행 제거
        if any(marker in raw_name for marker in ("합계", "총계", "소계", "TOTAL", "Total")):
            continue

        belong = _cell(row, belong_col)
        note = _cell(row, note_col)

        names = split_names(raw_name) or [raw_name]
        attendance = Attendance.ABSENT if ("불참" in note or "미참" in note) else Attendance.PRESENT
        amount = parse_amount(_cell(row, amount_col)) if amount_col is not None else 0

        adult_tickets = _int_cell(row, adult_col)
        child_tickets = _int_cell(row, child_col)
        if adult_col is None:
            adult_tickets = len(names) if attendance is Attendance.PRESENT else 0
        elif attendance is Attendance.ABSENT:
            adult_tickets = 0

        guest = Guest(
            name=format_display_name(names),
            names=names,
            amount=amount,
            relation=guess_relation(belong, note, raw_name),
            attendance=attendance,
            payment=Payment.coerce(note) if source is Source.EXCEL else Payment.TRANSFER,
            adult_tickets=adult_tickets,
            child_tickets=child_tickets,
            belong=belong,
            note=note,
            raw=" ".join(part for part in (raw_name, str(amount), belong, note) if part).strip(),
            source=source,
            guest_id=offset,
        )
        if amount <= 0:
            guest.add_warning(WARN_NO_AMOUNT)
        if not raw_name.strip():
            guest.add_warning(WARN_NO_NAME)
        guests.append(guest)

    return renumber(guests)


def parse_spreadsheet(file_path: str | Path, source: Source = Source.EXCEL) -> list[Guest]:
    """엑셀/CSV 파일을 읽어 Guest 목록을 만든다."""
    path = Path(file_path)
    if not path.exists():
        raise ExcelParseError(f"파일을 찾을 수 없습니다: {path}")
    rows = _load_rows(path)
    guests = parse_rows(rows, source=source)
    logger.info("스프레드시트 파싱 완료: %s (%d건)", path.name, len(guests))
    return guests
